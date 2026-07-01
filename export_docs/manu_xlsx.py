# -*- coding: utf-8 -*-
"""بروفورما تصنيع (Manufacturing Proforma) — مطابقة لقالب المصنع الفعلي (موديل/خشب/لون/يد/إكسسوار/Sağ-Sol)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

NAVY = "1E2B48"; GOLD = "C9A24B"; GOLD_BG = "FBF3E0"; LIGHT = "F5F7FA"; ROW = "FAFBFC"
WHITE = "FFFFFF"; GREY = "6B7177"; INK = "20242C"; STEEL = "27608B"

F_TITLE = Font(name="Georgia", size=15, bold=True, color=NAVY)
F_SUB = Font(name="Arial", size=8.5, color=GREY)
F_DOC = Font(name="Georgia", size=14, bold=True, color=NAVY)
F_BADGE = Font(name="Arial", size=8.5, bold=True, color="8A6D1C")
F_LBL = Font(name="Arial", size=8.5, bold=True, color=STEEL)
F_K = Font(name="Arial", size=9.5, color=GREY)
F_V = Font(name="Arial", size=10, bold=True, color=NAVY)
F_H = Font(name="Arial", size=8.5, bold=True, color=WHITE)
F_TXT = Font(name="Arial", size=9.5, color=INK)
F_BOLD = Font(name="Arial", size=10, bold=True, color=NAVY)
F_GRAND = Font(name="Georgia", size=12, bold=True, color=GOLD)
F_AUTH = Font(name="Arial", size=9, bold=True, color="7A611F")

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_ROW = PatternFill("solid", fgColor=ROW)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_GOLD_BG = PatternFill("solid", fgColor=GOLD_BG)

CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color="E1E4E8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
gold_thin = Side(style="thin", color=GOLD)
GOLD_BORDER = Border(top=gold_thin, bottom=gold_thin, left=gold_thin, right=gold_thin)

HEADERS = ["#", "Model / Code", "Wood", "Metal Color", "Pull Handle", "Accessory Color",
           "Sağ / Right", "Sol / Left", "Total Pcs", "Unit Price", "Total Price"]
NCOLS = len(HEADERS)


def _f(n):
    try: return f"{float(n):,.2f}"
    except Exception: return n

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def _initials(co):
    src = (co["short_name"] or co["name"] or "").strip()
    words = [w for w in src.split() if w]
    if not words: return "—"
    return words[0][:2].upper() if len(words) == 1 else (words[0][0] + words[1][0]).upper()

def _kv(ws, r, label, value, label_col=1, val_span=3):
    if not value: return
    ws.cell(row=r, column=label_col, value=label).font = F_K
    _merge(ws, r, label_col + 1, r, label_col + val_span)
    c = ws.cell(row=r, column=label_col + 1, value=value); c.font = F_TXT; c.alignment = LFT

def _banner(ws, r, ncols, text, font, fill=None, border=None, height=20):
    _merge(ws, r, 1, r, ncols)
    c = ws.cell(row=r, column=1, value=text); c.font = font; c.alignment = CTR
    if fill: c.fill = fill
    if border:
        for col in range(1, ncols + 1):
            ws.cell(row=r, column=col).border = border
    ws.row_dimensions[r].height = height


def build(order, company, customer, items):
    """order: sqlite3.Row من جدول manu_orders. company/customer: sqlite3.Row."""
    wb = Workbook(); ws = wb.active; ws.title = "Manufacturing Proforma"
    ws.sheet_view.showGridLines = False
    widths = [5, 16, 20, 18, 16, 18, 10, 10, 10, 11, 13]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    exporter_name = company["name"] or "Exporter"
    _merge(ws, 1, 1, 1, 6); ws.cell(row=1, column=1, value=exporter_name).font = F_TITLE
    ws.row_dimensions[1].height = 22
    _merge(ws, 2, 1, 2, 6); ws.cell(row=2, column=1, value="Steel Doors · Manufacturing Order").font = F_SUB
    _merge(ws, 1, 8, 1, NCOLS)
    tc = ws.cell(row=1, column=8, value="PROFORMA INVOICE (MANUFACTURING)"); tc.font = F_DOC; tc.alignment = RGT
    _merge(ws, 2, 8, 2, NCOLS)
    bc = ws.cell(row=2, column=8, value=f"REF {order['ref_no']}"); bc.font = F_BADGE; bc.alignment = RGT
    for col in range(1, NCOLS + 1):
        ws.cell(row=3, column=col).fill = FILL_GOLD
    ws.row_dimensions[3].height = 3

    r = 5
    _kv(ws, r, "TO", customer["name"], val_span=4); r += 1
    if customer["contact"]:
        _kv(ws, r, "CONTACT", customer["contact"], val_span=4); r += 1
    if customer["address"]:
        _kv(ws, r, "ADDRESS", customer["address"], val_span=4); r += 1
    _kv(ws, r, "PROFORMA NO", order["ref_no"], val_span=4); r += 1
    _kv(ws, r, "DATE", order["created_at"], val_span=4); r += 1
    _kv(ws, r, "DELIVERY", order["delivery"] or "—", val_span=4); r += 1
    _kv(ws, r, "PAYMENT TERM", order["payment_term"] or "—", val_span=4); r += 1
    _kv(ws, r, "COUNTRY OF ORIGIN", "TURKEY", val_span=4); r += 1
    _kv(ws, r, "COUNTRY OF MANUFACTURE", "TURKEY", val_span=4); r += 2

    r = _table_header(ws, r)
    cur = order["currency"] or "USD"
    tot_r = tot_l = tot_qty = tot_amt = 0
    for i, it in enumerate(items):
        qty = it.get("qty") or (it.get("right", 0) + it.get("left", 0))
        amt = qty * it.get("unit_price", 0)
        tot_r += it.get("right", 0); tot_l += it.get("left", 0); tot_qty += qty; tot_amt += amt
        vals = [i + 1, it.get("model", ""), it.get("wood", ""), it.get("metal_color", ""),
                it.get("handle", ""), it.get("accessory_color", ""), it.get("right", 0) or "",
                it.get("left", 0) or "", int(qty), _f(it.get("unit_price", 0)), _f(amt)]
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=j + 1, value=v); c.border = BORDER; c.font = F_TXT
            c.alignment = LFT if j == 1 else CTR
            if i % 2: c.fill = FILL_ROW
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = F_BOLD
    for col, val in [(7, tot_r or ""), (8, tot_l or ""), (9, int(tot_qty))]:
        c = ws.cell(row=r, column=col, value=val); c.font = F_BOLD; c.alignment = CTR
    _merge(ws, r, 10, r, 11)
    gc = ws.cell(row=r, column=10, value=f"{cur} {_f(tot_amt)}"); gc.font = F_GRAND; gc.fill = FILL_NAVY; gc.alignment = CTR
    ws.row_dimensions[r].height = 24
    r += 2

    if order["notes"]:
        _banner(ws, r, NCOLS, order["notes"], F_TXT, height=20); r += 2

    _banner(ws, r, NCOLS, "BENEFICIARY BANK DETAILS", F_LBL, fill=FILL_GOLD_BG, height=18); r += 1
    branch = (company["branch"] or "") + (f" (Code {company['branch_code']})" if company["branch_code"] else "")
    bank_rows = [("Company Name", company["name"]), ("Bank Name", company["bank_name"]), ("Branch", branch)]
    try:
        if company["account_no"]:
            bank_rows.append(("Account No", company["account_no"]))
    except Exception:
        pass
    bank_rows += [("IBAN No", company["iban"]), ("SWIFT Code", company["swift"])]
    for k, v in bank_rows:
        if v: _kv(ws, r, k, v, val_span=NCOLS - 3); r += 1
    r += 1
    _banner(ws, r, NCOLS, "THIS PROFORMA IS AUTHENTIC AND APPROVED BY THE EXPORTER",
            F_AUTH, fill=FILL_GOLD_BG, border=GOLD_BORDER, height=22)
    return wb


def _table_header(ws, r):
    for i, h in enumerate(HEADERS):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = F_H; c.fill = FILL_NAVY; c.alignment = CTR; c.border = BORDER
    ws.row_dimensions[r].height = 26
    return r + 1


def to_bytes(wb):
    bio = BytesIO(); wb.save(bio); return bio.getvalue()
