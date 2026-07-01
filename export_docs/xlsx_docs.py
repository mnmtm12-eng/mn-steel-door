# -*- coding: utf-8 -*-
"""توليد وثائق Excel احترافية (xlsx) لكل نوع وثيقة تصدير — بهوية الشركة المصدّرة."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

NAVY = "1E2B48"; STEEL = "27608B"; GOLD = "C9A24B"; GOLD_BG = "FBF3E0"
LIGHT = "F5F7FA"; ROW = "FAFBFC"; WHITE = "FFFFFF"; GREY = "6B7177"; INK = "20242C"

F_TITLE = Font(name="Georgia", size=17, bold=True, color=NAVY)
F_SUB = Font(name="Arial", size=8.5, color=GREY)
F_DOC = Font(name="Georgia", size=15, bold=True, color=NAVY)
F_BADGE = Font(name="Arial", size=8.5, bold=True, color="8A6D1C")
F_H = Font(name="Arial", size=9.5, bold=True, color=WHITE)
F_LBL = Font(name="Arial", size=8.5, bold=True, color=STEEL)
F_K = Font(name="Arial", size=9.5, color=GREY)
F_V = Font(name="Arial", size=10.5, bold=True, color=NAVY)
F_TXT = Font(name="Arial", size=10, color=INK)
F_BOLD = Font(name="Arial", size=10.5, bold=True, color=NAVY)
F_GRAND = Font(name="Georgia", size=13, bold=True, color=GOLD)
F_AUTH = Font(name="Arial", size=9.5, bold=True, color="7A611F")

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_ROW = PatternFill("solid", fgColor=ROW)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_GOLD_BG = PatternFill("solid", fgColor=GOLD_BG)

CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color="E1E4E8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
gold_thin = Side(style="thin", color=GOLD)
GOLD_BORDER = Border(top=gold_thin, bottom=gold_thin, left=gold_thin, right=gold_thin)


def _f(n):
    try: return f"{float(n):,.2f}"
    except Exception: return n

def _initials(co):
    src = (co["short_name"] or co["name"] or "").strip()
    words = [w for w in src.split() if w]
    if not words: return "—"
    return words[0][:2].upper() if len(words) == 1 else (words[0][0] + words[1][0]).upper()

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def _kv_row(ws, r, ncols, label, value, label_span=1, font_v=None):
    """صفّ (تسمية مدمجة | قيمة مدمجة) — يحلّ مشكلة قصّ النصوص نهائياً."""
    if not value: return
    _merge(ws, r, 1, r, max(1, label_span))
    lc = ws.cell(row=r, column=1, value=label); lc.font = F_K
    vc1 = label_span + 1
    if vc1 <= ncols:
        _merge(ws, r, vc1, r, ncols)
        vc = ws.cell(row=r, column=vc1, value=value)
        vc.font = font_v or F_TXT; vc.alignment = LFT

def _banner(ws, r, ncols, text, font, fill=None, border=None, height=22):
    _merge(ws, r, 1, r, ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = font; c.alignment = CTR
    if fill: c.fill = fill
    if border:
        for col in range(1, ncols + 1):
            ws.cell(row=r, column=col).border = border
    ws.row_dimensions[r].height = height

def _hdr(ws, ctx, ncols=5):
    s, co, cu, doc = ctx["s"], ctx["company"], ctx["cust"], ctx["doc"]
    exporter_name = co["name"] or "Exporter"
    # letterhead: exporter identity (لا يوضع اسم الموقع، بل الشركة المصدّرة فعلياً)
    _merge(ws, 1, 1, 1, max(2, ncols - 2))
    ws.cell(row=1, column=1, value=exporter_name).font = F_TITLE
    ws.row_dimensions[1].height = 24
    _merge(ws, 2, 1, 2, max(2, ncols - 2))
    ws.cell(row=2, column=1, value="Steel Doors · Export Documentation").font = F_SUB

    _merge(ws, 1, max(3, ncols - 1), 1, ncols)
    tc = ws.cell(row=1, column=max(3, ncols - 1), value=doc["label"]); tc.font = F_DOC; tc.alignment = RGT
    _merge(ws, 2, max(3, ncols - 1), 2, ncols)
    bc = ws.cell(row=2, column=max(3, ncols - 1), value=f"REF {s['ref_no']}"); bc.font = F_BADGE; bc.alignment = RGT

    # gold accent divider
    for col in range(1, ncols + 1):
        ws.cell(row=3, column=col).fill = FILL_GOLD
    ws.row_dimensions[3].height = 3

    r = 5
    pairs = [("REFERENCE", s["ref_no"]), ("DATE", s["created_at"]), ("HS CODE", ctx["hs"]), ("ORIGIN", ctx["origin"])]
    span = max(1, ncols // 4)
    col = 1
    for k, v in pairs:
        end = min(col + span - 1, ncols)
        _merge(ws, r, col, r, end); _merge(ws, r + 1, col, r + 1, end)
        ws.cell(row=r, column=col, value=k).font = F_LBL
        ws.cell(row=r + 1, column=col, value=v).font = F_V
        col = end + 1
        if col > ncols: break
    r += 3

    # exporter / consignee boxes
    half = max(1, ncols // 2)
    _merge(ws, r, 1, r, half)
    ws.cell(row=r, column=1, value="EXPORTER").font = F_LBL
    _merge(ws, r + 1, 1, r + 1, half)
    ws.cell(row=r + 1, column=1, value=exporter_name).font = F_BOLD
    ex = [x for x in [co["address"], (("Tel: " + co["phone"]) if co["phone"] else ""),
          (("Email: " + co["email"]) if co["email"] else ""),
          (("Tax No (VKN): " + co["tax_no"]) if (("tax_no" in co.keys()) and co["tax_no"]) else "")] if x]
    for i, line in enumerate(ex):
        _merge(ws, r + 2 + i, 1, r + 2 + i, half)
        c = ws.cell(row=r + 2 + i, column=1, value=line); c.font = F_TXT; c.alignment = LFT

    midc = half + 1
    _merge(ws, r, midc, r, ncols)
    ws.cell(row=r, column=midc, value="CONSIGNEE").font = F_LBL
    _merge(ws, r + 1, midc, r + 1, ncols)
    ws.cell(row=r + 1, column=midc, value=cu["name"]).font = F_BOLD
    cn = [x for x in [cu["contact"], cu["address"], (("Tel: " + cu["phone"]) if cu["phone"] else ""),
          (("Tax No: " + cu["tax_no"]) if cu["tax_no"] else "")] if x]
    for i, line in enumerate(cn):
        _merge(ws, r + 2 + i, midc, r + 2 + i, ncols)
        c = ws.cell(row=r + 2 + i, column=midc, value=line); c.font = F_TXT; c.alignment = LFT
    r += 2 + max(len(ex), len(cn)) + 1

    # shipment strip
    ship = [("Incoterm", s["incoterm"]), ("Loading", s["port_loading"] or "—"),
            ("Discharge", s["port_discharge"] or "—"), ("ACID No", s["acid_no"] or "—"),
            ("Country", cu["country"] or "—")]
    n = min(len(ship), ncols)
    base = ncols // n; extra = ncols % n; col = 1
    for i, (k, v) in enumerate(ship[:n]):
        w = base + (1 if i < extra else 0)
        end = min(col + w - 1, ncols)
        _merge(ws, r, col, r, end); _merge(ws, r + 1, col, r + 1, end)
        kc = ws.cell(row=r, column=col, value=k.upper()); kc.font = F_LBL; kc.fill = FILL_LIGHT
        vc = ws.cell(row=r + 1, column=col, value=v); vc.font = F_V; vc.fill = FILL_LIGHT
        for cc in range(col, end + 1):
            ws.cell(row=r, column=cc).fill = FILL_LIGHT
            ws.cell(row=r + 1, column=cc).fill = FILL_LIGHT
        col = end + 1
    return r + 3

def _table_header(ws, r, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i + 1, value=h)
        c.font = F_H; c.fill = FILL_NAVY; c.alignment = CTR; c.border = BORDER
    ws.row_dimensions[r].height = 20
    return r + 1

def _widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def build(doctype, ctx):
    s, co, cu = ctx["s"], ctx["company"], ctx["cust"]
    items, containers = ctx["items"], ctx["containers"]
    cur = s["currency"] or "USD"
    wb = Workbook(); ws = wb.active; ws.title = ctx["doc"]["label"][:31]
    ws.sheet_view.showGridLines = False

    if doctype in ("commercial", "proforma"):
        ncols = 9
        _widths(ws, [6, 24, 10, 10, 10, 11, 9, 13, 15]); r = _hdr(ws, ctx, ncols)
        r = _table_header(ws, r, ["#", "Model / Code", "Width (cm)", "Height (cm)", "Frame (cm)",
                                  "Opening", "Qty (pcs)", f"Unit Price ({cur})", f"Total ({cur})"])
        for i, it in enumerate(items):
            lt = it.get("line_total", it["qty"] * it["unit_price"])
            vals = [i + 1, it["name"], it.get("width", ""), it.get("height", ""), it.get("beam", ""),
                    it.get("opening", ""), int(it["qty"]), _f(it["unit_price"]), _f(lt)]
            for j, v in enumerate(vals):
                c = ws.cell(row=r, column=j + 1, value=v); c.border = BORDER
                c.alignment = LFT if j == 1 else CTR; c.font = F_TXT
                if i % 2: c.fill = FILL_ROW
            r += 1
        ws.cell(row=r, column=6, value="TOTAL").font = F_BOLD
        tq = ws.cell(row=r, column=7, value=s["qty_total"]); tq.font = F_BOLD; tq.alignment = CTR
        _merge(ws, r, 8, r, 9)
        gc = ws.cell(row=r, column=8, value=f"{cur} {_f(s['total'])}"); gc.font = F_GRAND; gc.fill = FILL_NAVY; gc.alignment = CTR
        ws.row_dimensions[r].height = 26
        r += 2

        if s["payment_terms"]:
            _kv_row(ws, r, ncols, "PAYMENT TERMS", s["payment_terms"], label_span=2); r += 1
        if co["bank_name"] or co["iban"] or co["swift"]:
            r += 1
            _banner(ws, r, ncols, "BENEFICIARY BANK DETAILS", F_LBL, fill=FILL_GOLD_BG, height=18); r += 1
            branch = co["branch"] + (f" (Code {co['branch_code']})" if co["branch_code"] else "")
            for k, v in [("Beneficiary", co["name"]), ("Bank", co["bank_name"]), ("Branch", branch),
                         ("IBAN", co["iban"]), ("SWIFT", co["swift"])]:
                if v: _kv_row(ws, r, ncols, k, v, label_span=2); r += 1
        r += 1
        _banner(ws, r, ncols, f"THIS {ctx['doc']['label'].upper()} IS AUTHENTIC AND APPROVED BY THE EXPORTER",
                F_AUTH, fill=FILL_GOLD_BG, border=GOLD_BORDER, height=22)

    elif doctype == "packing":
        ncols = 5
        _widths(ws, [7, 36, 20, 24, 20]); r = _hdr(ws, ctx, ncols)
        r = _table_header(ws, r, ["#", "Description of Goods", "Quantity", "", ""])
        for i, it in enumerate(items):
            for j, v in enumerate([i + 1, f"{it['name']}  ({it['width']}×{it['height']}×{it['beam']} cm)", f"{int(it['qty'])} pcs", "", ""]):
                c = ws.cell(row=r, column=j + 1, value=v); c.border = BORDER; c.font = F_TXT
                c.alignment = LFT if j == 1 else CTR
                if i % 2: c.fill = FILL_ROW
            r += 1
        _merge(ws, r, 2, r, 2)
        ws.cell(row=r, column=2, value="TOTAL").font = F_BOLD
        ws.cell(row=r, column=3, value=f"{s['qty_total']} pcs").font = F_BOLD; ws.cell(row=r, column=3).alignment = CTR
        r += 2
        _banner(ws, r, ncols, "CONTAINER / PACKING DETAILS", F_LBL, fill=FILL_GOLD_BG, height=18); r += 1
        r = _table_header(ws, r, ["#", "Container No.", "Seal No.", "Content", "Gross Weight (KG)"])
        for i, ct in enumerate(containers):
            for j, v in enumerate([i + 1, ct.get("no", "—") or "—", ct.get("seal", "—") or "—", ct.get("content", "—") or "—", _f(ct.get("weight", 0))]):
                c = ws.cell(row=r, column=j + 1, value=v); c.border = BORDER; c.font = F_TXT
                c.alignment = LFT if j == 3 else CTR
                if i % 2: c.fill = FILL_ROW
            r += 1
        if containers:
            ws.cell(row=r, column=4, value="TOTAL GROSS WEIGHT").font = F_BOLD
            gc = ws.cell(row=r, column=5, value=f"{_f(s['gross_weight'] or 0)} KG"); gc.font = F_GRAND; gc.fill = FILL_NAVY; gc.alignment = CTR
            r += 2
        _banner(ws, r, ncols, "THIS PACKING LIST IS AUTHENTIC AND APPROVED BY THE EXPORTER",
                F_AUTH, fill=FILL_GOLD_BG, border=GOLD_BORDER, height=22)

    elif doctype == "origin":
        ncols = 5
        _widths(ws, [6, 40, 14, 16, 18]); r = _hdr(ws, ctx, ncols)
        _banner(ws, r, ncols, "CERTIFICATE OF ORIGIN  ·  شهادة منشأ", F_DOC, height=24); r += 2
        r = _table_header(ws, r, ["#", "Description of Goods", "Quantity", "HS Code", "Country of Origin"])
        for i, it in enumerate(items):
            for j, v in enumerate([i + 1, f"{it['name']}  ({it['width']}×{it['height']}×{it['beam']} cm)", f"{int(it['qty'])} pcs", ctx["hs"], ctx["origin"]]):
                c = ws.cell(row=r, column=j + 1, value=v); c.border = BORDER; c.font = F_TXT
                c.alignment = LFT if j == 1 else CTR
                if i % 2: c.fill = FILL_ROW
            r += 1
        ws.cell(row=r, column=2, value="TOTAL").font = F_BOLD
        ws.cell(row=r, column=3, value=f"{s['qty_total']} pcs").font = F_BOLD; ws.cell(row=r, column=3).alignment = CTR
        r += 2
        _banner(ws, r, ncols, f"We hereby certify that the goods described above are of {ctx['origin']} origin, "
                f"ref. {s['ref_no']}, consignee {cu['name']}.", F_TXT, height=30); r += 2
        _banner(ws, r, ncols, "THIS CERTIFICATE OF ORIGIN IS AUTHENTIC AND APPROVED BY THE EXPORTER",
                F_AUTH, fill=FILL_GOLD_BG, border=GOLD_BORDER, height=22)

    elif doctype == "shipping":
        ncols = 5
        _widths(ws, [8, 30, 16, 18, 22]); r = _hdr(ws, ctx, ncols)
        _banner(ws, r, ncols, "TO: SHIPPING LINE / FREIGHT FORWARDER", F_BOLD, height=18); r += 1
        _banner(ws, r, ncols, f"Please arrange shipment for reference {s['ref_no']} per the details below.", F_TXT, height=20); r += 2
        r = _table_header(ws, r, ["Commodity", "HS Code", "Quantity", "Containers", "Gross Weight"])
        for j, v in enumerate(["Steel Doors", ctx["hs"], f"{s['qty_total']} pcs", s["container_count"], f"{_f(s['gross_weight'] or 0)} KG"]):
            c = ws.cell(row=r, column=j + 1, value=v); c.border = BORDER; c.font = F_TXT; c.alignment = CTR
        r += 2
        if containers:
            r = _table_header(ws, r, ["#", "Container No.", "Seal No.", "Content", "Weight (KG)"])
            for i, ct in enumerate(containers):
                for j, v in enumerate([i + 1, ct.get("no", "—") or "—", ct.get("seal", "—") or "—", ct.get("content", "—") or "—", _f(ct.get("weight", 0))]):
                    c = ws.cell(row=r, column=j + 1, value=v); c.border = BORDER; c.font = F_TXT
                    c.alignment = LFT if j == 3 else CTR
                r += 1
            r += 1
        special = s["notes"] or ("Please notify the consignee upon arrival. Ensure ACID number appears on all documents "
                                  "for Egyptian customs (NAFEZA) where applicable.")
        if s["acid_no"]:
            special += f"  ACID No: {s['acid_no']}"
        _banner(ws, r, ncols, "SPECIAL INSTRUCTIONS", F_LBL, fill=FILL_GOLD_BG, height=18); r += 1
        _banner(ws, r, ncols, special, F_TXT, height=34)

    return wb


def to_bytes(wb):
    bio = BytesIO(); wb.save(bio); return bio.getvalue()
