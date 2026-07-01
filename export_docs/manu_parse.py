# -*- coding: utf-8 -*-
"""تحليل ذكي لطلبات التصنيع: من ملف Excel أو نص ملصوق (من سكرين شوت) → بنود جاهزة للمراجعة."""
import re, unicodedata

MODEL_KEYS = ["MODEL", "CODE NO", "CODE", "GOODS DESCRIPTION", "DESCRIPTION"]
WOOD_KEYS = ["AHŞAP", "AHSAP", "WOOD"]
METAL_KEYS = ["KASA RENK", "METAL COLOR", "KASA"]
HANDLE_KEYS = ["ÇEKME", "CEKME", "PULL HANDLE", "HANDLE"]
ACCESSORY_KEYS = ["AKSESUAR", "ACCESSORY"]
RIGHT_KEYS = ["SAĞ", "SAG", "RIGHT"]
LEFT_KEYS = ["SOL", "LEFT"]
TOTALPCS_KEYS = ["TOTAL PCS", "TOTAL PIECES", "QUANTITY", "QTY"]
PRICE_KEYS = ["PRICE", "FIYAT"]
TOTALPRICE_KEYS = ["TOTAL PRICE", "TOTAL USD", "AMOUNT", "TOPLAM"]

DEFAULT_ATTR = "As shown in image / Görüldüğü gibi"


def _norm(s):
    if s is None: return ""
    s = str(s).upper()
    s = unicodedata.normalize("NFKD", s)
    return s.strip()

def _num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(",", "").replace("%", "")
    try: return float(s)
    except Exception: return None

def _match_col(headers, keys, exclude=None):
    exclude = exclude or []
    for i, h in enumerate(headers):
        hn = _norm(h)
        if not hn: continue
        if any(e in hn for e in exclude): continue
        if any(k in hn for k in keys):
            return i
    return None


def parse_excel(path):
    """يحلّل ملف بروفورما تصنيع (نمط NEWA/RANGE/SAIROGULLARI) → قائمة بنود."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    best_rows, best_score, best_idx = None, 0, None
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True, max_row=min(ws.max_row, 60)))
        for idx, row in enumerate(rows[:25]):
            headers = list(row)
            score = 0
            if _match_col(headers, MODEL_KEYS) is not None: score += 1
            if _match_col(headers, RIGHT_KEYS) is not None: score += 1
            if _match_col(headers, LEFT_KEYS) is not None: score += 1
            if _match_col(headers, TOTALPCS_KEYS) is not None: score += 1
            if _match_col(headers, PRICE_KEYS) is not None: score += 1
            if score > best_score:
                best_score, best_idx, best_rows = score, idx, rows
    wb.close()
    if best_rows is None or best_score < 2:
        return [], "لم يتمكّن التحليل من التعرّف على أعمدة الجدول تلقائياً — أضف البنود يدوياً."

    headers = list(best_rows[best_idx])
    col_model = _match_col(headers, MODEL_KEYS)
    col_wood = _match_col(headers, WOOD_KEYS)
    col_metal = _match_col(headers, METAL_KEYS)
    col_handle = _match_col(headers, HANDLE_KEYS)
    col_acc = _match_col(headers, ACCESSORY_KEYS)
    col_right = _match_col(headers, RIGHT_KEYS)
    col_left = _match_col(headers, LEFT_KEYS)
    col_qty = _match_col(headers, TOTALPCS_KEYS)
    col_tprice = _match_col(headers, TOTALPRICE_KEYS)
    col_price = _match_col(headers, PRICE_KEYS, exclude=["TOTAL"])
    if col_price is None and col_tprice is not None:
        col_price = None  # سنشتقّه من الإجمالي إذا لزم

    items = []
    for row in best_rows[best_idx + 1:]:
        if row is None or all(c is None for c in row):
            break
        model_val = row[col_model] if col_model is not None and col_model < len(row) else None
        model_s = str(model_val).strip() if model_val is not None else ""
        if not model_s:
            continue
        if _norm(model_s).startswith("TOTAL"):
            break
        right = _num(row[col_right]) if col_right is not None and col_right < len(row) else None
        left = _num(row[col_left]) if col_left is not None and col_left < len(row) else None
        qty = _num(row[col_qty]) if col_qty is not None and col_qty < len(row) else None
        price = _num(row[col_price]) if col_price is not None and col_price < len(row) else None
        tprice = _num(row[col_tprice]) if col_tprice is not None and col_tprice < len(row) else None
        right = right or 0; left = left or 0
        if qty is None:
            qty = right + left if (right or left) else 0
        if price is None and tprice and qty:
            price = round(tprice / qty, 4)
        if qty == 0 and not price:
            continue
        items.append({
            "model": model_s,
            "wood": str(row[col_wood]).strip() if col_wood is not None and col_wood < len(row) and row[col_wood] else DEFAULT_ATTR,
            "metal_color": str(row[col_metal]).strip() if col_metal is not None and col_metal < len(row) and row[col_metal] else DEFAULT_ATTR,
            "handle": str(row[col_handle]).strip() if col_handle is not None and col_handle < len(row) and row[col_handle] else DEFAULT_ATTR,
            "accessory_color": str(row[col_acc]).strip() if col_acc is not None and col_acc < len(row) and row[col_acc] else DEFAULT_ATTR,
            "right": right, "left": left, "qty": qty, "unit_price": price or 0,
        })
    if not items:
        return [], "التعرّف على الأعمدة نجح لكن لم توجد بنود صالحة — أضف البنود يدوياً."
    return items, None


CODE_RE = re.compile(r"(?:code|كود)\s*[:#]?\s*(\d+)", re.I)
NUM_RE = re.compile(r"\d+(?:[.,]\d+)?")


def parse_text(text):
    """يحلّل نصاً ملصوقاً (منسوخاً يدوياً من سكرين شوت الطلب) إلى بنود مبدئية."""
    items = []
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line or len(line) < 2:
            continue
        code_m = CODE_RE.search(line)
        model_label = f"Code {code_m.group(1)}" if code_m else None
        line_for_nums = line[:code_m.start()] + line[code_m.end():] if code_m else line
        nums = [float(n.replace(",", ".")) for n in NUM_RE.findall(line_for_nums)]
        if not nums and not model_label:
            continue
        if model_label:
            label = model_label
        else:
            label = NUM_RE.sub(" ", line)
            label = re.sub(r"[|;,\-–—:•*]+", " ", label).strip()
            label = re.sub(r"\s{2,}", " ", label)
            if not label:
                label = "بدون اسم"
        right = left = qty = price = 0
        low = line.lower()
        has_r = any(k in low for k in ["يمين", "sağ", "sag", "right"])
        has_l = any(k in low for k in ["شمال", "يسار", "sol", "left"])
        if len(nums) == 1:
            qty = nums[0]
        elif len(nums) == 2:
            if has_r or has_l:
                right, left = (nums[0], nums[1]) if not (has_l and not has_r) else (nums[1], nums[0])
                qty = right + left
            else:
                qty, price = nums[0], nums[1]
        elif len(nums) == 3:
            right, left, price = nums; qty = right + left
        elif len(nums) >= 4:
            right, left, qty, price = nums[0], nums[1], nums[2], nums[3]
        items.append({
            "model": label[:60], "wood": DEFAULT_ATTR, "metal_color": DEFAULT_ATTR,
            "handle": DEFAULT_ATTR, "accessory_color": DEFAULT_ATTR,
            "right": right, "left": left, "qty": qty or (right + left), "unit_price": price,
        })
    if not items:
        return [], "لم يُستخرج أي بند من النص — تأكد من وجود أرقام (كميات) بجانب كل موديل."
    return items, None
